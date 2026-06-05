"""报告生成模块 — 生成HTML、Excel和文本格式的财务分析报告"""

import json
import os
import traceback
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd

from config import OUTPUT_DIR

# 中文字体配置
plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

OUTPUT_DIR_ABS = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), OUTPUT_DIR)
os.makedirs(OUTPUT_DIR_ABS, exist_ok=True)


def _fmt(val, unit: str = "亿", decimals: int = 2) -> str:
    """格式化金额"""
    if val is None:
        return "N/A"
    if unit == "亿":
        return f"{val / 1e8:,.{decimals}f}"
    elif unit == "万":
        return f"{val / 1e4:,.{decimals}f}"
    return f"{val:,.{decimals}f}"


def _pct(val) -> str:
    """格式化百分比"""
    if val is None:
        return "N/A"
    return f"{val * 100:.2f}%"


def generate_charts(summary: dict, ratios: dict, stock_code: str, stock_name: str) -> list:
    """生成图表并返回图表文件路径列表"""
    chart_paths = []
    ts_data = summary.get("_time_series", {})

    try:
        # 1. 杜邦分析分解图
        dupont = _chart_dupont(ratios, stock_name)
        if dupont:
            chart_paths.append(dupont)

        # 2. 财务健康度雷达图
        radar = _chart_radar(ratios, stock_name)
        if radar:
            chart_paths.append(radar)

        # 3. 收入/利润趋势图
        income_ts = ts_data.get("income")
        if income_ts is not None and not income_ts.empty:
            trend = _chart_revenue_profit_trend(income_ts, stock_name)
            if trend:
                chart_paths.append(trend)

        # 4. 本福特定律对比图
        benford_chart = _chart_benford_digits()
        if benford_chart:
            chart_paths.append(benford_chart)

    except Exception as e:
        print(f"  [图表] 生成图表时出错: {e}")

    return chart_paths


def _chart_dupont(ratios: dict, stock_name: str) -> str:
    """杜邦分析分解图"""
    roe = ratios.get("roe")
    nm = ratios.get("net_margin")
    at = ratios.get("total_asset_turnover")
    em = ratios.get("debt_to_equity")

    if not any([roe, nm, at, em]):
        return None

    fig, ax = plt.subplots(figsize=(8, 4))

    labels = ["ROE", "净利率", "资产周转率", "权益乘数-1\n(杠杆)"]
    values = [
        roe * 100 if roe else 0,
        nm * 100 if nm else 0,
        at * 100 if at else 0,
        (em * 100) if em else 0,
    ]
    colors = ["#2196F3", "#4CAF50", "#FF9800", "#f44336"]

    bars = ax.bar(labels, values, color=colors, width=0.5)
    ax.set_title(f"{stock_name} - 杜邦分析关键驱动因素", fontsize=14, fontweight="bold")
    ax.set_ylabel("百分比 (%)")

    for bar, val in zip(bars, values):
        if val != 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                    f"{val:.1f}%", ha="center", va="bottom", fontsize=11, fontweight="bold")

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()

    path = os.path.join(OUTPUT_DIR_ABS, "chart_dupont.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


def _chart_radar(ratios: dict, stock_name: str) -> str:
    """财务健康度雷达图"""
    breakdown = ratios.get("health_score_breakdown", {})
    if not breakdown:
        return None

    categories = ["盈利能力", "偿债能力", "营运能力", "成长能力", "现金流"]
    values = [
        breakdown.get("profitability", 0),
        breakdown.get("solvency", 0),
        breakdown.get("efficiency", 0),
        breakdown.get("growth", 0),
        breakdown.get("cashflow", 0),
    ]

    N = len(categories)
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    values += values[:1]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.fill(angles, values, color="#2196F3", alpha=0.25)
    ax.plot(angles, values, color="#2196F3", linewidth=2)

    # 添加参考线
    for level in [25, 50, 75]:
        ref = [level] * (N + 1)
        ax.plot(angles, ref, color="gray", linewidth=0.5, alpha=0.5, linestyle="--")

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, fontsize=11)
    ax.set_ylim(0, 100)
    ax.set_yticks([25, 50, 75])
    ax.set_yticklabels(["25", "50", "75"], fontsize=8, color="gray")
    ax.set_title(f"{stock_name} - 财务健康度评分\n(总分: {ratios.get('health_score', 'N/A')})",
                 fontsize=14, fontweight="bold", pad=20)

    plt.tight_layout()
    path = os.path.join(OUTPUT_DIR_ABS, "chart_radar.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


def _chart_revenue_profit_trend(income_df: pd.DataFrame, stock_name: str) -> str:
    """收入与利润趋势图"""
    if "report_date" not in income_df.columns:
        return None

    df = income_df.sort_values("report_date").copy()
    dates = df["report_date"].astype(str).tolist()
    if len(dates) > 8:
        step = len(dates) // 6
        tick_idx = list(range(0, len(dates), step))
        tick_labels = [dates[i][:10] if i < len(dates) else "" for i in tick_idx]
    else:
        tick_idx = list(range(len(dates)))
        tick_labels = [d[:10] for d in dates]

    fig, ax1 = plt.subplots(figsize=(10, 5))

    rev_col = "operating_revenue" if "operating_revenue" in df.columns else None
    profit_col = "net_profit" if "net_profit" in df.columns else None

    x = range(len(dates))

    if rev_col:
        rev = df[rev_col].values / 1e8
        ax1.bar(x, rev, color="#E3F2FD", label="营业收入(亿)", alpha=0.8)
    if profit_col:
        profit = df[profit_col].values / 1e8
        ax1.plot(x, profit, "o-", color="#f44336", linewidth=2, markersize=6, label="净利润(亿)")

    ax1.set_xticks(tick_idx)
    ax1.set_xticklabels(tick_labels, rotation=30, fontsize=8)
    ax1.set_title(f"{stock_name} - 营业收入与净利润趋势", fontsize=14, fontweight="bold")
    ax1.legend(loc="upper left")
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)

    plt.tight_layout()
    path = os.path.join(OUTPUT_DIR_ABS, "chart_trend.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


def _chart_benford_digits() -> str:
    """本福特定律理论分布图"""
    import math
    BENFORD_PROBS = {d: math.log10(1 + 1 / d) for d in range(1, 10)}

    digits = list(range(1, 10))
    probs = [BENFORD_PROBS[d] * 100 for d in digits]

    fig, ax = plt.subplots(figsize=(8, 4))
    bars = ax.bar(digits, probs, color=["#1565C0", "#1976D2", "#1E88E5", "#2196F3",
                                         "#42A5F5", "#64B5F6", "#90CAF9", "#BBDEFB", "#E3F2FD"])
    ax.set_title("本福特定律——首位数字理论概率分布", fontsize=14, fontweight="bold")
    ax.set_xlabel("首位数字")
    ax.set_ylabel("出现概率 (%)")
    ax.set_xticks(digits)

    for bar, p in zip(bars, probs):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                f"{p:.1f}%", ha="center", va="bottom", fontsize=9)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()

    path = os.path.join(OUTPUT_DIR_ABS, "chart_benford.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    return path


def generate_excel_report(
    stock_code: str,
    stock_name: str,
    summary: dict,
    ratios: dict,
    benford_results: list,
    altman_result: dict,
    audit_result: dict,
) -> str:
    """生成Excel格式的财务分析报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [报告] openpyxl未安装，跳过Excel生成")
        return None

    wb = Workbook()

    # === Sheet 1: 财务指标汇总 ===
    ws1 = wb.active
    ws1.title = "财务指标汇总"

    header_font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    sub_header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    value_font = Font(name="Microsoft YaHei", size=11)
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"{stock_name}（{stock_code}）财务分析报告"
    ws1["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1565C0")
    ws1["A2"] = f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    row = 4
    # 基本信息
    ws1[f"A{row}"] = "基本信息"
    ws1[f"A{row}"].font = header_font
    ws1[f"A{row}"].fill = header_fill
    for c in range(1, 5):
        ws1[f"{get_column_letter(c)}{row}"].fill = header_fill

    row += 1
    basic_fields = [
        ("股票代码", stock_code),
        ("股票简称", stock_name),
        ("综合健康度评分", f"{ratios.get('health_score', 'N/A')} / 100"),
    ]
    for label, val in basic_fields:
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = sub_header_font
        ws1[f"B{row}"] = str(val) if val is not None else "N/A"
        row += 1

    # 财务比率
    ratio_groups = [
        ("盈利能力", [
            ("毛利率", ratios.get("gross_margin"), _pct),
            ("净利率", ratios.get("net_margin"), _pct),
            ("营业利润率", ratios.get("operating_margin"), _pct),
            ("ROE（净资产收益率）", ratios.get("roe"), _pct),
            ("ROA（总资产收益率）", ratios.get("roa"), _pct),
            ("核心利润占比", ratios.get("core_profit_ratio"), _pct),
        ]),
        ("偿债能力", [
            ("资产负债率", ratios.get("debt_ratio"), _pct),
            ("权益乘数", ratios.get("debt_to_equity"), lambda x: f"{x:.2f}" if x else "N/A"),
            ("流动比率", ratios.get("current_ratio"), lambda x: f"{x:.2f}" if x else "N/A"),
            ("速动比率", ratios.get("quick_ratio"), lambda x: f"{x:.2f}" if x else "N/A"),
            ("有息负债率", ratios.get("interest_bearing_debt_ratio"), _pct),
        ]),
        ("营运能力", [
            ("总资产周转率", ratios.get("total_asset_turnover"), lambda x: f"{x:.2f}" if x else "N/A"),
            ("应收账款周转天数", ratios.get("ar_turnover_days"), lambda x: f"{x:.0f}天" if x else "N/A"),
            ("存货周转天数", ratios.get("inventory_turnover_days"), lambda x: f"{x:.0f}天" if x else "N/A"),
            ("现金循环周期", ratios.get("cash_conversion_cycle"), lambda x: f"{x:.0f}天" if x else "N/A"),
        ]),
        ("成长能力", [
            ("营收同比增长率", ratios.get("revenue_growth_yoy"), _pct),
            ("净利润同比增长率", ratios.get("profit_growth_yoy"), _pct),
            ("3年营收CAGR", ratios.get("revenue_cagr_3y"), _pct),
        ]),
        ("现金流质量", [
            ("经营现金流/净利润", ratios.get("ocf_to_profit"), lambda x: f"{x:.2f}" if x else "N/A"),
            ("经营现金流/营收", ratios.get("ocf_to_revenue"), _pct),
            ("自由现金流(亿)", ratios.get("free_cashflow"), lambda x: _fmt(x) if x else "N/A"),
        ]),
        ("风险指标", [
            ("商誉/净资产", ratios.get("goodwill_to_equity"), _pct),
            ("固定资产占比", ratios.get("fixed_asset_ratio"), _pct),
            ("投资收益/营业利润", ratios.get("investment_income_to_profit"), _pct),
            ("公允价值变动/营业利润", ratios.get("fair_value_to_profit"), _pct),
        ]),
    ]

    for group_name, metrics in ratio_groups:
        row += 1
        ws1[f"A{row}"] = group_name
        ws1[f"A{row}"].font = header_font
        ws1[f"A{row}"].fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
        for c in range(1, 5):
            ws1[f"{get_column_letter(c)}{row}"].fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
        row += 1

        for label, val, fmt_func in metrics:
            ws1[f"B{row}"] = label
            ws1[f"B{row}"].font = value_font
            ws1[f"C{row}"] = fmt_func(val) if val is not None else "N/A"
            ws1[f"C{row}"].font = value_font
            row += 1

    # === Sheet 2: 本福特定律检验 ===
    ws2 = wb.create_sheet("本福特定律检验")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "本福特定律首位数字检验结果"
    ws2["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)

    headers = ["科目", "样本量", "卡方统计量", "P值", "是否异常", "MAD", "一致性"]
    for i, h in enumerate(headers, 1):
        ws2[f"{get_column_letter(i)}3"] = h
        ws2[f"{get_column_letter(i)}3"].font = Font(name="Microsoft YaHei", bold=True)

    for r, br in enumerate(benford_results, 4):
        ws2[f"A{r}"] = br.get("label", "")
        ws2[f"B{r}"] = br.get("sample_size", "N/A")
        ws2[f"C{r}"] = br.get("chi2_statistic", "N/A")
        ws2[f"D{r}"] = br.get("p_value", "N/A")
        ws2[f"E{r}"] = "⚠ 异常" if br.get("is_abnormal") else "✓ 正常"
        ws2[f"F{r}"] = br.get("mad", "N/A")
        ws2[f"G{r}"] = br.get("conformity_cn", "N/A")

        if br.get("is_abnormal"):
            for c in range(1, 8):
                ws2[f"{get_column_letter(c)}{r}"].fill = warn_fill

    # === Sheet 3: 审计发现 ===
    ws3 = wb.create_sheet("审计分析")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "审计角度分析发现"
    ws3["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)

    row = 3
    for finding in audit_result.get("findings", []):
        ws3[f"A{row}"] = f"【{finding['area']}】"
        ws3[f"A{row}"].font = Font(name="Microsoft YaHei", size=12, bold=True,
                                   color="FF5722" if finding["risk_level"] == "high" else "FF9800")
        row += 1
        ws3[f"A{row}"] = finding["description"]
        row += 1
        for d in finding.get("details", []):
            ws3[f"B{row}"] = f"• {d}"
            row += 1
        row += 1

    # === Sheet 4: 核心财务数据 ===
    ws4 = wb.create_sheet("核心财务数据")
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "最新一期核心财务数据（单位：亿元）"
    ws4["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)

    data_fields = [
        ("营业收入", summary.get("operating_revenue")),
        ("营业成本", summary.get("operating_cost")),
        ("销售费用", summary.get("selling_expenses")),
        ("管理费用", summary.get("admin_expenses")),
        ("研发费用", summary.get("rd_expenses")),
        ("财务费用", summary.get("finance_expenses")),
        ("营业利润", summary.get("operating_profit")),
        ("利润总额", summary.get("total_profit")),
        ("净利润", summary.get("net_profit")),
        ("归母净利润", summary.get("net_profit_parent")),
        ("总资产", summary.get("total_assets")),
        ("总负债", summary.get("total_liabilities")),
        ("净资产", summary.get("total_equity")),
        ("经营活动现金流", summary.get("operating_cashflow")),
        ("投资活动现金流", summary.get("investing_cashflow")),
        ("筹资活动现金流", summary.get("financing_cashflow")),
    ]

    for r, (label, val) in enumerate(data_fields, 3):
        ws4[f"B{r}"] = label
        ws4[f"C{r}"] = _fmt(val) if val is not None else "N/A"

    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4]:
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15

    filepath = os.path.join(OUTPUT_DIR_ABS, f"{stock_code}_{stock_name}_财务分析报告.xlsx")
    wb.save(filepath)
    return filepath


def generate_html_report(
    stock_code: str,
    stock_name: str,
    basic_info: dict,
    summary: dict,
    ratios: dict,
    benford_results: list,
    altman_result: dict,
    audit_result: dict,
    chart_paths: list,
    anomaly_results: dict = None,
) -> str:
    """生成HTML格式的财务分析报告"""
    template_path = os.path.join(os.path.dirname(__file__), "templates", "report.html")
    if not os.path.exists(template_path):
        template_path = os.path.join(os.path.dirname(__file__), "templates", "report_template.html")

    health_score = ratios.get("health_score", "N/A")
    health_breakdown = ratios.get("health_score_breakdown", {})

    # 风险颜色
    if isinstance(health_score, (int, float)):
        if health_score >= 70:
            score_color = "#4CAF50"
            score_label = "良好"
        elif health_score >= 50:
            score_color = "#FF9800"
            score_label = "一般"
        else:
            score_color = "#f44336"
            score_label = "需关注"
    else:
        score_color = "#9E9E9E"
        score_label = "N/A"

    # 构建HTML（内联方式，不依赖外部模板）
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{stock_name}({stock_code}) 财务分析报告</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: "Microsoft YaHei", "SimHei", sans-serif; background: #f5f7fa; color: #333; line-height: 1.8; }}
.container {{ max-width: 1100px; margin: 0 auto; padding: 20px; }}
.header {{ background: linear-gradient(135deg, #1565C0, #1E88E5); color: white; padding: 40px; border-radius: 12px; margin-bottom: 24px; text-align: center; }}
.header h1 {{ font-size: 28px; margin-bottom: 8px; }}
.header .subtitle {{ opacity: 0.85; font-size: 14px; }}
.card {{ background: white; border-radius: 12px; padding: 28px; margin-bottom: 20px; box-shadow: 0 2px 12px rgba(0,0,0,0.06); }}
.card h2 {{ font-size: 20px; color: #1565C0; border-bottom: 2px solid #E3F2FD; padding-bottom: 12px; margin-bottom: 20px; }}
.card h3 {{ font-size: 16px; color: #424242; margin: 16px 0 10px 0; }}
.score-badge {{ display: inline-block; background: {score_color}; color: white; padding: 12px 32px; border-radius: 24px; font-size: 24px; font-weight: bold; margin: 12px 0; }}
.score-label {{ font-size: 14px; opacity: 0.9; margin-left: 8px; }}
table {{ width: 100%; border-collapse: collapse; margin: 12px 0; }}
th, td {{ padding: 10px 14px; text-align: left; border-bottom: 1px solid #e0e0e0; font-size: 14px; }}
th {{ background: #f5f5f5; font-weight: 600; color: #424242; }}
tr:hover {{ background: #fafafa; }}
.metric-good {{ color: #4CAF50; font-weight: bold; }}
.metric-warn {{ color: #FF9800; font-weight: bold; }}
.metric-bad {{ color: #f44336; font-weight: bold; }}
.finding {{ margin: 14px 0; padding: 14px; border-left: 4px solid #ddd; background: #fafafa; border-radius: 4px; }}
.finding.high {{ border-left-color: #f44336; background: #FFF5F5; }}
.finding.medium {{ border-left-color: #FF9800; background: #FFF8E1; }}
.finding.info {{ border-left-color: #2196F3; background: #E3F2FD; }}
.finding .area-title {{ font-weight: bold; font-size: 15px; margin-bottom: 6px; }}
.finding li {{ margin: 4px 0 4px 18px; font-size: 14px; }}
.chart-container {{ text-align: center; margin: 16px 0; }}
.chart-container img {{ max-width: 100%; border-radius: 8px; }}
.alert {{ padding: 14px; border-radius: 8px; margin: 10px 0; font-size: 14px; }}
.alert-danger {{ background: #FFEBEE; color: #C62828; border: 1px solid #FFCDD2; }}
.alert-warning {{ background: #FFF8E1; color: #F57F17; border: 1px solid #FFECB3; }}
.alert-info {{ background: #E3F2FD; color: #1565C0; border: 1px solid #BBDEFB; }}
.alert-success {{ background: #E8F5E9; color: #2E7D32; border: 1px solid #C8E6C9; }}
.footer {{ text-align: center; font-size: 12px; color: #9E9E9E; margin: 40px 0 20px 0; }}
.badge {{ display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; }}
.badge-red {{ background: #FFCDD2; color: #C62828; }}
.badge-yellow {{ background: #FFECB3; color: #F57F17; }}
.badge-green {{ background: #C8E6C9; color: #2E7D32; }}
.badge-blue {{ background: #BBDEFB; color: #1565C0; }}
</style>
</head>
<body>
<div class="container">

<div class="header">
    <h1>{stock_name}（{stock_code}）财务分析报告</h1>
    <div class="subtitle">报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 数据来源: AKShare</div>
</div>

<!-- 综合评分 -->
<div class="card" style="text-align: center;">
    <h2>综合财务健康度评分</h2>
    <div class="score-badge">{health_score}<span class="score-label">{score_label}</span></div>
    <table style="max-width: 600px; margin: 20px auto;">
        <tr>
            <th>盈利能力</th><td>{health_breakdown.get('profitability', 'N/A')}</td>
            <th>偿债能力</th><td>{health_breakdown.get('solvency', 'N/A')}</td>
        </tr>
        <tr>
            <th>营运能力</th><td>{health_breakdown.get('efficiency', 'N/A')}</td>
            <th>成长能力</th><td>{health_breakdown.get('growth', 'N/A')}</td>
        </tr>
        <tr>
            <th>现金流质量</th><td>{health_breakdown.get('cashflow', 'N/A')}</td>
            <th></th><td></td>
        </tr>
    </table>
</div>

<!-- 一、财务指标 -->
<div class="card">
    <h2>一、核心财务指标</h2>
"""

    # 盈利能力
    html += "<h3>1.1 盈利能力</h3><table>"
    html += "<tr><th>指标</th><th>数值</th><th>评价</th></tr>"
    profit_metrics = [
        ("毛利率", ratios.get("gross_margin"), ">30%为优"),
        ("净利率", ratios.get("net_margin"), ">10%为优"),
        ("ROE", ratios.get("roe"), ">15%为优"),
        ("ROA", ratios.get("roa"), ">5%为优"),
        ("核心利润占比", ratios.get("core_profit_ratio"), ">80%为优"),
    ]
    for name, val, note in profit_metrics:
        cls = ""
        if val is not None:
            if name == "ROE" and val > 0.15:
                cls = "metric-good"
            elif name == "ROE" and val < 0.05:
                cls = "metric-bad"
            elif name == "毛利率" and val > 0.3:
                cls = "metric-good"
        html += f"<tr><td>{name}</td><td class='{cls}'>{_pct(val)}</td><td style='color:#999;font-size:12px;'>{note}</td></tr>"
    html += "</table>"

    # 偿债能力
    html += "<h3>1.2 偿债能力</h3><table>"
    html += "<tr><th>指标</th><th>数值</th><th>参考区间</th></tr>"
    solvency_metrics = [
        ("资产负债率", ratios.get("debt_ratio"), "40%-60%"),
        ("流动比率", ratios.get("current_ratio"), ">1.5"),
        ("速动比率", ratios.get("quick_ratio"), ">0.8"),
        ("有息负债率", ratios.get("interest_bearing_debt_ratio"), "<30%"),
    ]
    for name, val, ref in solvency_metrics:
        html += f"<tr><td>{name}</td><td>{_pct(val)}</td><td style='color:#999;font-size:12px;'>{ref}</td></tr>"
    html += "</table>"

    # 营运能力
    html += "<h3>1.3 营运能力</h3><table>"
    html += "<tr><th>指标</th><th>数值</th></tr>"
    efficiency_metrics = [
        ("总资产周转率", f"{ratios.get('total_asset_turnover'):.2f}次" if ratios.get("total_asset_turnover") else "N/A"),
        ("应收账款周转天数", f"{ratios.get('ar_turnover_days'):.0f}天" if ratios.get("ar_turnover_days") else "N/A"),
        ("存货周转天数", f"{ratios.get('inventory_turnover_days'):.0f}天" if ratios.get("inventory_turnover_days") else "N/A"),
        ("现金循环周期", f"{ratios.get('cash_conversion_cycle'):.0f}天" if ratios.get("cash_conversion_cycle") else "N/A"),
    ]
    for name, val in efficiency_metrics:
        html += f"<tr><td>{name}</td><td>{val}</td></tr>"
    html += "</table>"

    # 成长能力
    html += "<h3>1.4 成长能力</h3><table>"
    html += "<tr><th>指标</th><th>数值</th></tr>"
    growth_metrics = [
        ("营收同比增长率", _pct(ratios.get("revenue_growth_yoy"))),
        ("净利润同比增长率", _pct(ratios.get("profit_growth_yoy"))),
        ("3年营收CAGR", _pct(ratios.get("revenue_cagr_3y"))),
    ]
    for name, val in growth_metrics:
        html += f"<tr><td>{name}</td><td>{val}</td></tr>"
    html += "</table>"

    # 现金流
    html += "<h3>1.5 现金流质量</h3><table>"
    html += "<tr><th>指标</th><th>数值</th><th>说明</th></tr>"
    cf_metrics = [
        ("经营现金流/净利润", f"{ratios.get('ocf_to_profit'):.2f}" if ratios.get("ocf_to_profit") is not None else "N/A", ">1为优质"),
        ("经营现金流/营收", _pct(ratios.get("ocf_to_revenue")), ">10%为优"),
        ("自由现金流", f"{ratios.get('free_cashflow')/1e8:,.2f}亿" if ratios.get("free_cashflow") else "N/A", "持续为正佳"),
    ]
    for name, val, note in cf_metrics:
        html += f"<tr><td>{name}</td><td>{val}</td><td style='color:#999;font-size:12px;'>{note}</td></tr>"
    html += "</table>"

    # 杜邦分析
    html += "<h3>1.6 杜邦分析（ROE驱动因素分解）</h3><table>"
    html += "<tr><th>驱动因素</th><th>数值</th><th>含义</th></tr>"
    html += f"<tr><td>净利率</td><td>{_pct(ratios.get('net_margin'))}</td><td>每元收入创造多少利润</td></tr>"
    html += f"<tr><td>资产周转率</td><td>{ratios.get('total_asset_turnover'):.2f}次</td><td>资产使用效率</td></tr>"
    html += f"<tr><td>权益乘数</td><td>{ratios.get('debt_to_equity'):.2f}</td><td>财务杠杆水平</td></tr>"
    html += f"<tr style='font-weight:bold;background:#f5f5f5;'><td>ROE</td><td>{_pct(ratios.get('roe'))}</td><td>股东回报率</td></tr>"
    html += "</table>"

    html += "</div>\n"

    # 二、值得注意的财报内容
    html += "<div class='card'><h2>二、财务报告值得注意的内容</h2>\n"

    # 核心财务数据一览
    html += "<h3>2.1 最新一期核心数据</h3><table>"
    html += "<tr><th>项目(亿元)</th><th>数值</th><th>项目(亿元)</th><th>数值</th></tr>"
    data_pairs = [
        ("营业收入", _fmt(summary.get("operating_revenue")), "营业成本", _fmt(summary.get("operating_cost"))),
        ("营业利润", _fmt(summary.get("operating_profit")), "利润总额", _fmt(summary.get("total_profit"))),
        ("净利润", _fmt(summary.get("net_profit")), "归母净利润", _fmt(summary.get("net_profit_parent"))),
        ("总资产", _fmt(summary.get("total_assets")), "净资产", _fmt(summary.get("total_equity"))),
        ("总负债", _fmt(summary.get("total_liabilities")), "资产负债率", _pct(ratios.get("debt_ratio"))),
        ("经营现金流", _fmt(summary.get("operating_cashflow")), "自由现金流", _fmt(ratios.get("free_cashflow"))),
    ]
    for a, av, b, bv in data_pairs:
        html += f"<tr><td>{a}</td><td><b>{av}</b></td><td>{b}</td><td><b>{bv}</b></td></tr>"
    html += "</table>"

    # 异常数据提示
    html += "<h3>2.2 数据异常提示</h3>"
    warnings_list = []

    # 费用结构异常
    selling_ratio = ratios.get("selling_expense_ratio")
    admin_ratio = ratios.get("admin_expense_ratio")
    if selling_ratio and selling_ratio > 0.30:
        warnings_list.append(("warning", f"销售费用率高达{_pct(selling_ratio)}，销售驱动的业务模式需关注效率"))
    if admin_ratio and admin_ratio > 0.15:
        warnings_list.append(("warning", f"管理费用率偏高({_pct(admin_ratio)})，管理效率有待提升"))

    # 商誉风险
    gw_to_eq = ratios.get("goodwill_to_equity")
    if gw_to_eq and gw_to_eq > 0.30:
        warnings_list.append(("danger", f"商誉占净资产{_pct(gw_to_eq)}，远超20%警戒线，减值风险极高"))
    elif gw_to_eq and gw_to_eq > 0.10:
        warnings_list.append(("warning", f"商誉占净资产{_pct(gw_to_eq)}，需关注减值风险"))

    # 利润与现金流背离
    ocf = summary.get("operating_cashflow")
    np_val = summary.get("net_profit")
    if ocf and np_val and np_val > 0 and ocf < np_val * 0.3:
        warnings_list.append(("danger", f"经营现金流仅覆盖净利润的{ocf/np_val:.1%}，利润含金量严重不足"))
    elif ocf and np_val and np_val > 0 and ocf < 0:
        warnings_list.append(("danger", "净利润为正但经营现金流为负，利润质量存在重大疑问"))

    # 投资/公允价值占比
    inv_to_profit = ratios.get("investment_income_to_profit")
    fv_to_profit = ratios.get("fair_value_to_profit")
    if inv_to_profit and abs(inv_to_profit) > 0.50:
        warnings_list.append(("warning", f"投资收益占营业利润{_pct(inv_to_profit)}，主营业务盈利能力存疑"))
    if fv_to_profit and abs(fv_to_profit) > 0.30:
        warnings_list.append(("warning", f"公允价值变动占营业利润{_pct(fv_to_profit)}，利润波动性较大"))

    # 减值损失
    asset_imp = summary.get("asset_impairment_loss") or 0
    credit_imp = summary.get("credit_impairment_loss") or 0
    if abs(asset_imp) + abs(credit_imp) > 1e8:
        warnings_list.append(("info", f"当期确认资产减值{abs(asset_imp)/1e8:.2f}亿、信用减值{abs(credit_imp)/1e8:.2f}亿"))

    if warnings_list:
        for level, msg in warnings_list:
            cls = {"danger": "alert-danger", "warning": "alert-warning", "info": "alert-info"}.get(level, "alert-info")
            html += f"<div class='alert {cls}'>{msg}</div>"
    else:
        html += "<div class='alert alert-success'>未发现显著异常数据</div>"

    html += "</div>\n"

    # 三、财务角度分析
    html += "<div class='card'><h2>三、财务角度分析</h2>\n"

    # 综合分析文字
    html += "<h3>3.1 综合分析</h3>"

    # 自动生成分析文字
    analysis_texts = []

    # 盈利能力判断
    roe = ratios.get("roe")
    gross_m = ratios.get("gross_margin")
    net_m = ratios.get("net_margin")
    if roe and roe > 0.15:
        analysis_texts.append(f"公司ROE为{_pct(roe)}，处于较优水平，反映股东回报能力较强。")
    elif roe and roe > 0.05:
        analysis_texts.append(f"公司ROE为{_pct(roe)}，处于中等水平，股东回报能力一般。")
    elif roe is not None:
        analysis_texts.append(f"公司ROE仅{_pct(roe)}，低于资本成本，股东价值创造能力不足。")

    if gross_m and gross_m > 0.40:
        analysis_texts.append(f"毛利率{_pct(gross_m)}，产品竞争力较强，具有较好的定价权。")
    elif gross_m and gross_m < 0.15:
        analysis_texts.append(f"毛利率仅{_pct(gross_m)}，产品或服务附加值较低，成本压力较大。")

    # 偿债能力判断
    debt_r = ratios.get("debt_ratio")
    if debt_r and debt_r > 0.70:
        analysis_texts.append(f"资产负债率高达{_pct(debt_r)}，财务杠杆较高，偿债压力值得关注。")
    elif debt_r and debt_r < 0.30:
        analysis_texts.append(f"资产负债率仅{_pct(debt_r)}，财务杠杆利用不足，可能损失了税盾收益。")

    # 现金流判断
    ocf_ratio = ratios.get("ocf_to_profit")
    if ocf_ratio is not None and ocf_ratio > 1.0:
        analysis_texts.append("经营现金流充分覆盖净利润，利润含金量高，盈利质量优秀。")
    elif ocf_ratio is not None and ocf_ratio < 0:
        analysis_texts.append(f"⚠ 经营现金流与净利润严重背离（OCF/NP={ocf_ratio:.2f}），利润质量存在重大疑问。")

    # 成长性判断
    rev_g = ratios.get("revenue_growth_yoy")
    if rev_g is not None and rev_g > 0.20:
        analysis_texts.append(f"营收同比增长{_pct(rev_g)}，处于快速增长阶段。")
    elif rev_g is not None and rev_g < 0:
        analysis_texts.append(f"⚠ 营收同比下滑{_pct(rev_g)}，业务收缩风险需关注。")

    for text in analysis_texts:
        html += f"<p style='margin:8px 0;text-indent:2em;'>{text}</p>"

    html += "</div>\n"

    # 四、审计角度分析
    html += "<div class='card'><h2>四、审计角度分析</h2>\n"

    # Altman Z-Score
    if altman_result and "error" not in altman_result:
        z = altman_result["z_score"]
        risk = altman_result["risk_level"]
        bc = {"green": "badge-green", "yellow": "badge-yellow", "red": "badge-red"}.get(
            altman_result.get("risk_color", ""), "badge-blue")
        html += f"""
        <h3>4.1 Altman Z-Score 破产风险评估</h3>
        <p>Z-Score: <b>{z}</b> <span class='badge {bc}'>{risk}</span>（{altman_result.get('version', '')}）</p>
        <table>
        <tr><th>指标</th><th>数值</th></tr>
        """
        for k, v in altman_result.get("components", {}).items():
            html += f"<tr><td>{k}</td><td>{v}</td></tr>"
        html += "</table>"

    # 审计发现
    html += "<h3>4.2 审计关注事项</h3>"
    for finding in audit_result.get("findings", []):
        level = finding["risk_level"]
        area = finding["area"]
        desc = finding["description"]
        details = finding.get("details", [])
        html += f"<div class='finding {level}'>"
        html += f"<div class='area-title'>{area}</div>"
        html += f"<p>{desc}</p>"
        if details:
            html += "<ul>"
            for d in details:
                html += f"<li>{d}</li>"
            html += "</ul>"
        html += "</div>"

    # 本福特定律结果
    html += "<h3>4.3 本福特定律数据质量检验</h3>"
    html += """<p style='font-size:13px;color:#666;margin-bottom:12px;'>本福特定律通过分析大量财务数据首位数字的分布规律来检测是否存在人为篡改。若实际分布显著偏离理论分布（p<0.05），可能存在数据异常。</p>"""
    html += "<table><tr><th>科目</th><th>样本量</th><th>P值</th><th>MAD</th><th>一致性</th><th>结果</th></tr>"
    for br in benford_results:
        if "error" in br:
            html += f"<tr><td>{br.get('label', '')}</td><td colspan='5' style='color:#999;'>{br['error']}</td></tr>"
        else:
            is_ab = br.get("is_abnormal", False)
            badge_html = '<span class="badge badge-red">异常</span>' if is_ab else '<span class="badge badge-green">正常</span>'
            html += f"<tr><td>{br.get('label', '')}</td><td>{br.get('sample_size', '')}</td><td>{br.get('p_value', '')}</td><td>{br.get('mad', '')}</td><td>{br.get('conformity_cn', '')}</td><td>{badge_html}</td></tr>"
    html += "</table>"

    # 关键审计事项预判
    html += "<h3>4.4 会计师视角总结</h3>"
    total_risks = audit_result.get("total_risks", 0)
    high_risks = audit_result.get("high_risks", 0)
    if high_risks > 0:
        html += f"<div class='alert alert-danger'>从审计角度看，本报告期存在 <b>{high_risks}</b> 个高风险领域和 <b>{total_risks}</b> 个中高风险领域，建议重点关注。</div>"
    elif total_risks > 0:
        html += f"<div class='alert alert-warning'>从审计角度看，本报告期存在 <b>{total_risks}</b> 个需要关注的风险领域。</div>"
    else:
        html += "<div class='alert alert-success'>从审计角度看，未发现重大风险信号。</div>"

    html += "</div>\n"

    # 图表
    if chart_paths:
        html += "<div class='card'><h2>五、分析图表</h2>"
        for cp in chart_paths:
            fname = os.path.basename(cp)
            html += f"<div class='chart-container'><img src='{fname}' alt='{fname}'></div>"
        html += "</div>\n"

    # 免责声明
    html += f"""
<div class='footer'>
    <p>免责声明：本报告由自动化财务分析系统生成，仅供研究参考，<b>不构成任何投资建议</b>。</p>
    <p>数据来源：公开市场数据（AKShare）| 分析日期：{datetime.now().strftime('%Y-%m-%d')}</p>
</div>

</div>
</body>
</html>"""

    filepath = os.path.join(OUTPUT_DIR_ABS, f"{stock_code}_{stock_name}_财务分析报告.html")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)

    return filepath
